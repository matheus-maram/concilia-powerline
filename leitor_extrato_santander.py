# leitor_extrato_santander_xlsx.py
import re
import pandas as pd
from openpyxl import load_workbook
import io

def _split_tipo_responsavel(historico):
    """
    Divide histórico em (Tipo Movimento, Responsável) usando 2+ espaços como separador.
    """
    if not isinstance(historico, str):
        return historico, None
    h = historico.strip()
    if "  " in h:
        partes = re.split(r"\s{2,}", h, maxsplit=1)
        tipo = partes[0].strip()
        resp = partes[1].strip() if len(partes) > 1 else None
        return tipo, resp
    return h, None


def ler_extrato_santander_xlsx(file) -> pd.DataFrame:
    """
    Lê extrato Santander em XLSX e retorna DataFrame estruturado.
    Aceita tanto caminho (str) quanto UploadedFile (Streamlit).
    """
    # Se vier do Streamlit, UploadedFile tem atributo .read()
    if not isinstance(file, (str, bytes)):
        file.seek(0)  # garante que está no início
        raw = pd.read_excel(file, header=None, engine="openpyxl")
    else:
        raw = pd.read_excel(file, header=None, engine="openpyxl")

    # Agência e Conta na primeira linha
    agencia = str(raw.iat[0, 1]).strip() if raw.shape[1] > 1 else None
    conta   = str(raw.iat[0, 3]).strip() if raw.shape[1] > 3 else None

    # Dados a partir da linha 4 (índice 3)
    df = raw.iloc[3:, :5].copy()
    df.columns = ["Data", "Vazio", "Historico", "Documento", "Valor"]

    # Limpeza
    df = df.dropna(how="all")
    df = df.loc[~df["Historico"].astype(str).str.upper().str.contains("SALDO", na=False)]
    df["Data"] = pd.to_datetime(df["Data"], errors="coerce", dayfirst=True)
    df = df[df["Data"].notna()].copy()

    # Documento como texto
    df["Documento"] = df["Documento"].astype(str).str.strip().replace({"nan": None, "None": None})

    # Tipo Movimento e Responsável
    tipo_resp = df["Historico"].apply(_split_tipo_responsavel)
    df["Tipo Movimento"] = tipo_resp.apply(lambda x: x[0])
    df["Responsável"] = tipo_resp.apply(lambda x: x[1])

    # Garantir que Valor é numérico
    df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce")

    # Entrada ou Saída
    def classificar_fluxo(v):
        if pd.isna(v):
            return None
        if v > 0:
            return "Entrada"
        elif v < 0:
            return "Saída"
        return "Neutro"

    df["Tipo de Fluxo"] = df["Valor"].apply(classificar_fluxo)

    # DataFrame final
    df_final = df.assign(**{
        "Agência": agencia,
        "Conta": conta
    })[["Agência", "Conta", "Data", "Tipo Movimento", "Responsável", "Documento", "Valor", "Tipo de Fluxo"]].reset_index(drop=True)

    return df_final


if __name__ == "__main__":
    caminho = "exportar-Santander - Extrato 11 de setembro de 2025-4591-130106767.xlsx"
    extrato = ler_extrato_santander_xlsx(caminho)

    # Exportar Excel
    output_path = "extrato_limpo.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        extrato.to_excel(writer, index=False, sheet_name="Extrato")

    # Reabrir para aplicar formatação
    wb = load_workbook(output_path)
    ws = wb["Extrato"]

    # Formatar colunas específicas
    for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
        col_letter = col[0].column_letter
        if col_letter == "C":  # Data
            for cell in col:
                cell.number_format = "DD/MM/YYYY"
        if col_letter == "G":  # Valor
            for cell in col:
                cell.number_format = u'R$ #.##0,00'  # formato brasileiro

    # Ajustar largura automática
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_path)

    # CSV simples
    extrato.to_csv("extrato_limpo.csv", index=False, encoding="utf-8-sig")

    print(f"✅ Extrato processado e formatado! Registros: {len(extrato)}")
    print(extrato.head())
